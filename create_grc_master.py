from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
wb.remove(ws)

sheets = {
  "Risk_Register": ["Risk_ID","Risk_Statement","Asset","Threat","Vulnerability","Impact","Likelihood","Rating","Treatment","Evidence/Notes"],
  "Control_Matrix": ["Control_ID","Control","Objective","Implementation_Point","Evidence","Frequency","Owner","Status"],
  "Evidence_Tracker": ["Evidence_ID","Control","Method","Frequency","Owner","Link/Path","Notes"],
  "Issues_MAP": ["Issue_ID","Finding","Severity","Root_Cause","Recommendation","Owner","Due_Date","Status","Verification_Evidence"],
  "Vendor_Register": ["Vendor","Service","Risk","Review_Cadence","Owner","Open_Items","Notes"],
  "Obligations_Register": ["Obligation","Source","Applies_to","Summary","Evidence","Status"],
  "Metrics": ["Metric","Definition","Target","Current","Trend","Cadence"]
}

for name, headers in sheets.items():
    ws = wb.create_sheet(name)
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(16, min(42, len(h)+4))

# seed minimal rows consistent with SDG repo analysis
wb["Risk_Register"].append(["R-01","OTP setup returns raw secret (demo risk)","MFA","Attacker/insider","Secret exposure at setup","5","2","10","Mitigate","auth.py /auth/otp/setup"])
wb["Risk_Register"].append(["R-02","seed_demo hard-coded credentials","SDLC","Attacker","Credential reuse risk","4","3","12","Mitigate","seed_demo.py"])
wb["Risk_Register"].append(["R-03","Full traceback stored in query_runs.error_message","Logging","Attacker/insider","Sensitive leakage via logs","4","3","12","Mitigate","gateway.py"])

wb.save("artifacts/11_GRC_Tooling/GRC_Master.xlsx")
print("Created artifacts/11_GRC_Tooling/GRC_Master.xlsx")
